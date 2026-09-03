"""Shared styling and cover-block helpers for the AVSAR document set.

Every document in docs/ is generated from these helpers so that the cover block,
the colour palette and the table styling stay identical across the ten documents.

    python docs/tools/generate_all.py
"""

from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT = "AVSAR"
PROJECT_LONG = "AVSAR - Startup-Friendly Public Procurement Platform"
VERSION = "1.0"
DOC_DATE = "03 September 2026"
PREPARED_BY = "TandSol"
CLASSIFICATION = "Internal - Smart India Hackathon submission"

DOC_IDS = {
    "handbook": ("AVSAR-HBK-001", "Project Handbook (README.md)"),
    "charter": ("AVSAR-CHR-002", "Project Charter"),
    "pmp": ("AVSAR-PMP-003", "Project Management Plan"),
    "rtm": ("AVSAR-RTM-004", "Requirements Traceability Matrix"),
    "sdd": ("AVSAR-SDD-005", "Software Design Document"),
    "code": ("AVSAR-CDR-006", "Code Register"),
    "risk": ("AVSAR-RSK-007", "Risk Register"),
    "timeline": ("AVSAR-TML-008", "Project Timeline"),
    "dpdp": ("AVSAR-DPD-009", "DPDP Compliance Tracker"),
    "retention": ("AVSAR-RET-010", "Data Retention Policy"),
    "qab": ("AVSAR-QAB-011", "SIH Question Bank"),
}

# Palette mirrors client/src/styles/tokens.css so documents and UI agree.
NAVY = "0B2447"
NAVY_MID = "19406F"
ACCENT = "05639E"
SAFFRON = "C26A0D"
GREEN = "067647"
AMBER = "B54708"
RED = "B42318"
INK = "344054"
GREY_BG = "F2F4F7"
LIGHT_BG = "F9FAFB"
BORDER = "D0D5DD"

THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------- workbook


def header_font():
    return Font(name="Calibri", size=10, bold=True, color="FFFFFF")


def body_font(bold=False, color=INK, size=10):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)


def write_table(ws, headers, rows, start_row=1, widths=None, wrap_cols=(), freeze=True):
    """Write a header row plus data rows with consistent banding and borders."""
    for c, head in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=head)
        cell.font = header_font()
        cell.fill = fill(NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[start_row].height = 30

    for r, row in enumerate(rows, start=start_row + 1):
        banded = (r - start_row) % 2 == 0
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body_font()
            cell.border = BOX
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(c in wrap_cols),
            )
            if banded:
                cell.fill = fill(LIGHT_BG)

    if widths:
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    )
    return start_row + len(rows)


def cover_sheet(wb, key, purpose, revisions, extra=None):
    """Standard cover tab: document control block, purpose, revision history."""
    doc_id, title = DOC_IDS[key]
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96

    ws["A1"] = PROJECT_LONG
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = title
    ws["A2"].font = Font(name="Calibri", size=13, color=ACCENT)
    ws.merge_cells("A1:B1")
    ws.merge_cells("A2:B2")

    control = [
        ("Document ID", doc_id),
        ("Version", VERSION),
        ("Date", DOC_DATE),
        ("Prepared by", PREPARED_BY),
        ("Status", "Living"),
        ("Related documents", ", ".join(v[0] for k, v in DOC_IDS.items() if k != key)),
        ("Classification", CLASSIFICATION),
    ]
    row = 4
    for label, value in control:
        ws.cell(row=row, column=1, value=label).font = body_font(bold=True, color=NAVY)
        ws.cell(row=row, column=1).fill = fill(GREY_BG)
        ws.cell(row=row, column=1).border = BOX
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = body_font()
        cell.border = BOX
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Purpose").font = body_font(bold=True, color=NAVY)
    cell = ws.cell(row=row, column=2, value=purpose)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.font = body_font()
    ws.row_dimensions[row].height = 58
    row += 2

    ws.cell(row=row, column=1, value="Revision history").font = body_font(bold=True, color=NAVY)
    row += 1
    write_table(
        ws,
        ["Version", "Date", "Author", "Change"],
        revisions,
        start_row=row,
        widths=[26, 96],
        wrap_cols=(4,),
        freeze=False,
    )
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 74

    if extra:
        row += len(revisions) + 3
        for label, value in extra:
            ws.cell(row=row, column=1, value=label).font = body_font(bold=True, color=NAVY)
            c = ws.cell(row=row, column=2, value=value)
            c.font = body_font()
            c.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
    return ws


def legend_sheet(wb, blocks):
    """Legend tab: one titled block per code set."""
    ws = wb.create_sheet("Legend")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 96

    row = 1
    for title, pairs in blocks:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = fill(NAVY)
        ws.cell(row=row, column=2).fill = fill(NAVY)
        cell.border = BOX
        ws.cell(row=row, column=2).border = BOX
        row += 1
        for code, meaning in pairs:
            k = ws.cell(row=row, column=1, value=code)
            k.font = body_font(bold=True)
            k.border = BOX
            v = ws.cell(row=row, column=2, value=meaning)
            v.font = body_font()
            v.border = BOX
            v.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        row += 1
    return ws


def summary_sheet(wb, title, metrics, note=None):
    """Summary tab. Every value is a live formula, never a hardcoded count."""
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)

    row = 3
    for label, formula, comment in metrics:
        k = ws.cell(row=row, column=1, value=label)
        k.font = body_font(bold=True)
        k.border = BOX
        k.fill = fill(GREY_BG)
        v = ws.cell(row=row, column=2, value=formula)
        v.font = body_font()
        v.border = BOX
        v.alignment = Alignment(horizontal="center")
        c = ws.cell(row=row, column=3, value=comment)
        c.font = body_font(color="667085")
        c.border = BOX
        c.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    if note:
        row += 1
        cell = ws.cell(row=row, column=1, value=note)
        cell.font = body_font(color="667085")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    return ws


def new_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def add_status_validation(ws, col_letter, first_row, last_row, options):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
