"""AVSAR-RSK-007 Risk Register and AVSAR-TML-008 Project Timeline."""

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common import (
    BOX, GREEN, GREY_BG, NAVY, RED, SAFFRON, add_status_validation, body_font,
    cover_sheet, fill, legend_sheet, new_workbook, summary_sheet, write_table,
)

# ============================================================ risk register

RISK_HEADERS = ["Risk ID", "Category", "Description", "Cause / Trigger", "Impact",
                "Likelihood (1-5)", "Impact (1-5)", "Score", "Severity", "Response",
                "Mitigation", "Owner", "Status", "Target Date"]
RISK_WIDTHS = [9, 16, 58, 46, 46, 13, 12, 8, 11, 12, 62, 22, 12, 13]
RISK_WRAP = (3, 4, 5, 11)

# Risk ID, Category, Description, Cause, Impact, L, I, Response, Mitigation, Owner, Status, Target
RISKS = [
    ("R-01", "Adoption", "Departments do not publish problem statements because writing an outcome is harder than copying a specification.",
     "No in-house practice of outcome-based specification; no worked examples to copy.",
     "The platform has no supply of problems and the whole mechanism stalls at stage one.",
     4, 5, "Mitigate",
     "The authoring form makes baseline, KPI, unit and direction separate mandatory fields and refuses to publish without a measurable KPI. Ship a library of published problem statements as worked examples, and a one-page authoring guide in the reference section.",
     "Mission Director", "Open", "31 Oct 2026"),
    ("R-02", "Legal", "A single-source award after a pilot is challenged as favouritism.",
     "Single-source procurement is inherently contestable where the pilot participant is also the awardee.",
     "Award set aside; a chilling effect on every other department considering the route.",
     3, 5, "Mitigate",
     "Four controls in combination: the problem statement is public before anyone applies; eligibility is decided by statute not opinion; scoring is blind against a rubric published in advance with a mandatory conflict-of-interest declaration and a locked score; and every procurement carries a written justification and a named GFR provision on its face. Where more than one pilot clears the gate the mode is a limited tender among participants under Rule 162, not single source.",
     "Director (Procurement)", "Open", "Continuous"),
    ("R-03", "Compliance", "Personal data handled during a pilot is processed beyond the declared purpose.",
     "A startup gains access to citizen data in a live environment for a limited operational purpose.",
     "DPDP Act 2023 breach, penalty exposure, and reputational damage to the mechanism itself.",
     3, 5, "Mitigate",
     "Pilot activation is blocked by the platform until the data processing agreement is recorded (HTTP 412, asserted in the end-to-end suite). Purpose is bounded by the pilot scope; the retention policy sets an erasure trigger per data category; the audit trail records every access-granting transition.",
     "Nodal Officer", "Open", "Continuous"),
    ("R-04", "Financial", "A pilot consumes budget and produces no usable outcome.",
     "Genuine technical uncertainty, which is the reason a pilot exists.",
     "Sunk cost and, more damaging, a department that concludes the whole model does not work.",
     4, 3, "Accept",
     "Accepted by design and priced in: the pilot budget is capped and ring-fenced, payment is released only against accepted milestone evidence, and the agreement carries an exit clause. A failed pilot is published on the transparency board rather than hidden, so the base rate is visible and a single failure is not read as a verdict on the mechanism.",
     "Department Head", "Open", "Continuous"),
    ("R-05", "Integrity", "An evaluator scores an applicant with whom they have an undeclared interest.",
     "Small expert pools in specialised sectors mean genuine overlap is common.",
     "A tainted shortlist that cannot be defended, and a grievance that succeeds.",
     3, 4, "Mitigate",
     "A conflict-of-interest declaration is a precondition the API enforces before any score is accepted. The first pass is blind. Scores lock on submission. Committee dispersion above twenty marks forces a reconciliation sitting. Every submission is written to the hash-chained audit trail with the evaluator identified.",
     "Nodal Officer", "Open", "Continuous"),
    ("R-06", "Financial", "Payment to a startup breaches the statutory forty-five-day window.",
     "Treasury and PFMS processing queues sit outside the platform's control.",
     "Interest liability under the MSMED Act, and cash-flow failure at a supplier for whom this is existential.",
     4, 4, "Mitigate",
     "The clock starts automatically on milestone acceptance and is visible to both parties. Breaches are counted on the public transparency board. The startup has a named grievance category with a fifteen-day SLA that escalates to the department head.",
     "Director (Procurement)", "Open", "Continuous"),
    ("R-07", "Technical", "A solution proven at pilot scale fails at production scale.",
     "A pilot is deliberately small; some failure modes appear only at volume.",
     "A live public service degrades, and the rate contract has already been extended to other departments.",
     3, 4, "Mitigate",
     "Scalability and interoperability is a scored technical criterion. The problem statement declares the indicative scale-up volume up front so the applicant designs for it. Rate contracts run two years with an exit, and the registry records adoption counts so a systemic problem surfaces early.",
     "Chief Engineer", "Open", "Continuous"),
    ("R-08", "Adoption", "Startups do not apply because the pilot value does not justify the effort.",
     "A capped pilot is small relative to the cost of selling into government.",
     "Thin applicant pools; the department concludes the market cannot deliver.",
     3, 4, "Mitigate",
     "Every problem statement must declare the indicative scale-up value and volume, which is what makes a small pilot worth the effort. EMD and tender fee are exempt, so applying costs only the work. The Proven Solutions Registry means one cleared pilot can generate orders from several departments.",
     "Mission Director", "Open", "31 Dec 2026"),
    ("R-09", "Security", "The audit log is altered to conceal an irregular award.",
     "Privileged database access exists in any real deployment.",
     "Loss of the evidentiary basis for every decision on the platform; audit findings against the department.",
     2, 5, "Mitigate",
     "The log is append-only and hash-chained: each entry hashes its predecessor, so any retrospective edit invalidates every subsequent link. A full-chain verification is exposed to the administrator and to audit, and the end-to-end suite asserts integrity after a complete lifecycle run.",
     "Scientist-E (NIC)", "Open", "Continuous"),
    ("R-10", "Technical", "Startup identity claims are accepted without verification.",
     "The demonstration build stubs MCA, GSTN and Udyam verification.",
     "An ineligible entity clears the gate; every downstream decision is unsound.",
     4, 4, "Mitigate",
     "The gate already records the declaration, the verdict and the rule, so a false claim is attributable. Live API verification against MCA (CIN), GSTN (GSTIN) and the Udyam registry is the first item on the deployment roadmap and is a hard precondition for any live award.",
     "Scientist-E (NIC)", "Open", "30 Nov 2026"),
    ("R-11", "Compliance", "The interface fails a GIGW 3.0 or WCAG 2.1 AA conformance audit.",
     "Conformance was designed for but not independently tested.",
     "Cannot be deployed on a government domain; rework late in the programme.",
     3, 3, "Mitigate",
     "Built to the standard from the outset: semantic landmarks, skip link, focus never suppressed, no meaning carried by colour alone, reduced-motion honoured. A third-party conformance audit is scheduled before deployment and is recorded as a gate, not an aspiration.",
     "Scientist-E (NIC)", "Open", "31 Jan 2027"),
    ("R-12", "Operational", "Departments duplicate effort by publishing problem statements another department has already solved.",
     "No habit of checking what has already been proven elsewhere.",
     "The central benefit of the mechanism is lost and public money buys the same lesson twice.",
     3, 3, "Mitigate",
     "The Proven Solutions Registry is a first-class navigation item on the public site and in every departmental workspace, and it carries the measured pilot KPIs so a second department sees evidence rather than a claim. Adoption is one click and skips discovery, evaluation and pilot.",
     "Mission Director", "Open", "Continuous"),
    ("R-13", "Security", "A departmental account is compromised and used to sanction a procurement.",
     "Local password authentication in the demonstration build.",
     "Fraudulent award and payment.",
     2, 5, "Mitigate",
     "Sanction authority is reserved to the department head and cannot be exercised by a procurement officer. Every action is attributed on the audit trail. NIC single sign-on with second-factor authentication replaces local passwords before deployment; the administrator can suspend an account immediately and the session fails on the next request rather than at token expiry.",
     "Scientist-E (NIC)", "Open", "31 Dec 2026"),
    ("R-14", "Operational", "Evaluator capacity becomes the bottleneck as volume grows.",
     "A small pool of domain experts, each with a day job.",
     "Evaluation timelines slip and the cycle-time advantage over a conventional tender disappears.",
     4, 3, "Mitigate",
     "The assignment screen shows current open load per evaluator so work is spread deliberately. The rubric is published so scoring is bounded rather than open-ended. A national evaluator panel across departments is the structural answer and is in the phase-two scope.",
     "Nodal Officer", "Open", "31 Mar 2027"),
    ("R-15", "Legal", "Intellectual property terms deter the applicants the department most wants.",
     "A default of government ownership is common in procurement templates.",
     "The strongest teams decline to apply and the applicant pool selects for the least differentiated.",
     3, 4, "Mitigate",
     "Startup retention of intellectual property is the platform default, is displayed prominently on every problem statement, and is recorded in the pilot agreement. A department choosing joint or government ownership must select it explicitly, which makes the trade-off visible at authoring time.",
     "Director (Procurement)", "Open", "Continuous"),
    ("R-16", "Technical", "The platform cannot integrate with GeM, PFMS or CPPP within the programme timeline.",
     "Onboarding to central systems is a governance process, not an engineering task.",
     "Manual re-entry of orders and payment status; the single-file-number claim weakens.",
     3, 3, "Transfer",
     "Integration points are already isolated behind named fields (GeM contract id, PFMS reference, purchase order number) so a live adapter replaces a stub without schema change. Onboarding is initiated in parallel with development and tracked as a dependency owned outside the engineering team.",
     "Mission Director", "Open", "28 Feb 2027"),
]

# =============================================================== timeline

PHASES = [
    ("P1", "Discovery and policy mapping", "Map the statutory basis; confirm which GFR relaxations apply and what they require on the file. Interview procurement and nodal officers on where a startup bid actually fails today.",
     "01 Jul 2026", "18 Jul 2026", 14, "Policy lead", "Sequential"),
    ("P2", "Charter and requirements", "Write the Charter; derive and baseline the RTM; agree the five-stage model and the gate at each boundary.",
     "21 Jul 2026", "01 Aug 2026", 10, "Project manager", "Sequential"),
    ("P3", "Data model and workflow design", "Sixteen-table schema; declare every legal state transition in one module; design the eligibility gate so each check cites its rule.",
     "04 Aug 2026", "15 Aug 2026", 10, "Solution architect", "Sequential"),
    ("P4", "Design system", "Token layer, component layer, accessibility policy, dense-table and form patterns. Built in parallel with the API so screens are never blocked on styling.",
     "11 Aug 2026", "22 Aug 2026", 10, "Design lead", "Parallel"),
    ("P5", "API - identity and challenge", "Authentication, role authorisation, startup registration with the statutory gate, problem statement authoring, approval and publication.",
     "18 Aug 2026", "27 Aug 2026", 8, "Backend engineer", "Sequential"),
    ("P6", "API - evaluation and pilot", "Application submission through both gates, committee assignment, blind scoring with lock, pilot creation, DPDP precondition, milestones, KPI readings.",
     "25 Aug 2026", "03 Sep 2026", 8, "Backend engineer", "Parallel"),
    ("P7", "API - procurement and registry", "Evidence-gated procurement with GFR modes, purchase orders, payment ledger with the statutory clock, Proven Solutions Registry and cross-department adoption.",
     "28 Aug 2026", "05 Sep 2026", 7, "Backend engineer", "Parallel"),
    ("P8", "Audit and transparency", "Hash-chained audit trail and verification, public transparency board with the conversion funnel and cycle-time medians.",
     "01 Sep 2026", "08 Sep 2026", 6, "Backend engineer", "Parallel"),
    ("P9", "Client - public site", "Landing, reference, transparency board, public problem statements, startup registry, proven solutions, registration and sign-in.",
     "25 Aug 2026", "05 Sep 2026", 10, "Frontend engineer", "Parallel"),
    ("P10", "Client - role workspaces", "Seven role-aware workspaces across twenty-four routes: authoring, discovery, scoring, pilot monitoring, procurement, payments, grievance, audit and administration.",
     "01 Sep 2026", "12 Sep 2026", 10, "Frontend engineer", "Parallel"),
    ("P11", "Seed dataset", "A deterministic dataset that exercises every state in the workflow, so a reviewer sees the whole lifecycle without clicking through it.",
     "08 Sep 2026", "12 Sep 2026", 5, "Backend engineer", "Parallel"),
    ("P12", "End-to-end test suite", "Thirty-two assertions across all five stages and the cross-cutting controls, against a live API on a throwaway database.",
     "10 Sep 2026", "17 Sep 2026", 6, "QA lead", "Sequential"),
    ("P13", "Document set", "The ten-document PM-SETU set with a live traceability chain from Charter to Risk Register.",
     "15 Sep 2026", "24 Sep 2026", 8, "Project manager", "Parallel"),
    ("P14", "Accessibility and security pass", "Keyboard traversal, contrast and screen-reader review; dependency audit; secrets and error-disclosure review.",
     "18 Sep 2026", "25 Sep 2026", 6, "QA lead", "Parallel"),
    ("P15", "Hackathon submission and demonstration", "Demonstration script, question bank, rehearsal against each role, submission.",
     "25 Sep 2026", "30 Sep 2026", 4, "Project manager", "Sequential"),
    ("P16", "Deployment readiness (post-hackathon)", "Live MCA, GSTN and Udyam verification; NIC single sign-on; PostgreSQL; GeM, PFMS and CPPP onboarding; VAPT by a CERT-In empanelled auditor; third-party GIGW audit.",
     "01 Oct 2026", "31 Mar 2027", 130, "Mission Director", "Sequential"),
]

WEEKS = [
    ("W1", "01-07 Jul"), ("W2", "08-14 Jul"), ("W3", "15-21 Jul"), ("W4", "22-28 Jul"),
    ("W5", "29 Jul-04 Aug"), ("W6", "05-11 Aug"), ("W7", "12-18 Aug"), ("W8", "19-25 Aug"),
    ("W9", "26 Aug-01 Sep"), ("W10", "02-08 Sep"), ("W11", "09-15 Sep"), ("W12", "16-22 Sep"),
    ("W13", "23-29 Sep"), ("W14", "30 Sep-06 Oct"), ("Q4", "Oct-Dec 2026"), ("Q1", "Jan-Mar 2027"),
]

# Inclusive week-column span per phase, indexed from 1.
GANTT = {
    "P1": (1, 3), "P2": (4, 5), "P3": (6, 7), "P4": (7, 8), "P5": (8, 9),
    "P6": (9, 10), "P7": (9, 10), "P8": (10, 11), "P9": (9, 10), "P10": (10, 11),
    "P11": (11, 11), "P12": (11, 12), "P13": (12, 13), "P14": (12, 13),
    "P15": (13, 14), "P16": (15, 16),
}


def build_risk(path):
    wb = new_workbook()

    cover_sheet(
        wb, "risk",
        "Risks to the AVSAR mechanism and to its delivery, scored on likelihood and impact. Score "
        "and Severity are formulas, not typed values, so re-scoring a risk reclassifies it "
        "automatically. Each mitigation states the specific control that addresses the risk, not an "
        "intention.",
        [("1.0", "03 Sep 2026", "TandSol", "Baseline. Sixteen risks across adoption, legal, "
          "compliance, financial, integrity, security, technical and operational categories.")],
        extra=[("Scoring", "Score = Likelihood x Impact. High >= 15, Medium 8-14, Low 1-7. "
                           "Both columns are formulas; see the Scoring Key tab.")],
    )

    legend_sheet(wb, [
        ("Category", [
            ("Adoption", "Whether departments and startups actually use the mechanism"),
            ("Legal", "Challenge to an award, or to the process that produced it"),
            ("Compliance", "Statutory or regulatory obligation"),
            ("Financial", "Public money at risk, or supplier cash flow"),
            ("Integrity", "Fairness and defensibility of a decision"),
            ("Security", "Confidentiality, integrity or availability of the platform"),
            ("Technical", "Engineering or scaling risk"),
            ("Operational", "Capacity, process and coordination"),
        ]),
        ("Response strategy", [
            ("Avoid", "Change the approach so the risk cannot arise"),
            ("Mitigate", "Reduce likelihood or impact with a named control"),
            ("Transfer", "Move the exposure to a party better placed to carry it"),
            ("Accept", "Consciously carry it, with the reasoning recorded"),
        ]),
        ("Status", [
            ("Open", "Live, being managed"),
            ("Mitigated", "Control in place and evidenced; residual score recorded"),
            ("Closed", "No longer applicable; reason recorded"),
            ("Realised", "The risk occurred; managed as an issue"),
        ]),
        ("Severity bands", [
            ("High", "Score 15 to 25. Named owner, mitigation with a target date, reviewed weekly."),
            ("Medium", "Score 8 to 14. Mitigation planned, reviewed fortnightly."),
            ("Low", "Score 1 to 7. Monitored, reviewed at each phase gate."),
        ]),
    ])

    ws = wb.create_sheet("Risk Register")
    rows = []
    for r in RISKS:
        rid, cat, desc, cause, impact, likelihood, impact_score, response, mitigation, owner, status, target = r
        rows.append([rid, cat, desc, cause, impact, likelihood, impact_score, None, None,
                     response, mitigation, owner, status, target])
    write_table(ws, RISK_HEADERS, rows, widths=RISK_WIDTHS, wrap_cols=RISK_WRAP)

    # Score and Severity are formulas so a re-score reclassifies automatically.
    for i in range(len(rows)):
        r = i + 2
        ws.cell(row=r, column=8, value=f"=F{r}*G{r}").font = body_font(bold=True)
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=8).border = BOX
        sev = ws.cell(row=r, column=9, value=f'=IF(H{r}>=15,"High",IF(H{r}>=8,"Medium","Low"))')
        sev.font = body_font(bold=True)
        sev.alignment = Alignment(horizontal="center")
        sev.border = BOX

    last = len(rows) + 1
    for rng, colour, op, val in [
        (f"I2:I{last}", RED, "equal", '"High"'),
        (f"I2:I{last}", SAFFRON, "equal", '"Medium"'),
        (f"I2:I{last}", GREEN, "equal", '"Low"'),
    ]:
        ws.conditional_formatting.add(rng, CellIsRule(
            operator=op, formula=[val],
            font=Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=colour),
        ))
    add_status_validation(ws, "M", 2, last, ["Open", "Mitigated", "Closed", "Realised"])
    add_status_validation(ws, "J", 2, last, ["Avoid", "Mitigate", "Transfer", "Accept"])

    key = wb.create_sheet("Scoring Key")
    key.sheet_view.showGridLines = False
    key["A1"] = "Risk scoring key"
    key["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    key["A3"] = "Score = Likelihood x Impact"
    key["A3"].font = body_font(bold=True)
    key["A4"] = "High >= 15 (red)   Medium 8-14 (amber)   Low 1-7 (green)"
    key["A4"].font = body_font()

    key["A6"] = "Likelihood"
    key["A6"].font = body_font(bold=True, color=NAVY)
    for i, (n, meaning) in enumerate([
        (1, "Rare - no comparable occurrence in similar programmes"),
        (2, "Unlikely - has occurred elsewhere but conditions differ"),
        (3, "Possible - has occurred in comparable programmes"),
        (4, "Likely - expected unless actively managed"),
        (5, "Almost certain - will occur without a control"),
    ], start=7):
        key.cell(row=i, column=1, value=n).font = body_font(bold=True)
        key.cell(row=i, column=2, value=meaning).font = body_font()

    key["A13"] = "Impact"
    key["A13"].font = body_font(bold=True, color=NAVY)
    for i, (n, meaning) in enumerate([
        (1, "Negligible - absorbed without schedule or cost effect"),
        (2, "Minor - local rework, no external effect"),
        (3, "Moderate - schedule slip or a visible degradation of the mechanism"),
        (4, "Major - a stage of the mechanism stops working as intended"),
        (5, "Severe - award set aside, statutory breach, or loss of confidence in the model"),
    ], start=14):
        key.cell(row=i, column=1, value=n).font = body_font(bold=True)
        key.cell(row=i, column=2, value=meaning).font = body_font()

    key["A21"] = "Matrix (Likelihood down, Impact across)"
    key["A21"].font = body_font(bold=True, color=NAVY)
    for c in range(1, 6):
        h = key.cell(row=22, column=c + 1, value=c)
        h.font = body_font(bold=True, color="FFFFFF")
        h.fill = fill(NAVY)
        h.alignment = Alignment(horizontal="center")
    for l in range(1, 6):
        rh = key.cell(row=22 + l, column=1, value=l)
        rh.font = body_font(bold=True, color="FFFFFF")
        rh.fill = fill(NAVY)
        rh.alignment = Alignment(horizontal="center")
        for c in range(1, 6):
            score = l * c
            cell = key.cell(row=22 + l, column=c + 1, value=score)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BOX
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = fill(RED if score >= 15 else SAFFRON if score >= 8 else GREEN)
    key.column_dimensions["A"].width = 14
    key.column_dimensions["B"].width = 76

    summary_sheet(wb, "Risk register - summary", [
        ("Risks registered", f"=COUNTA('Risk Register'!A2:A{last})", "Total rows"),
        ("High severity", f"=COUNTIF('Risk Register'!I2:I{last},\"High\")", "Score 15 or above; weekly review"),
        ("Medium severity", f"=COUNTIF('Risk Register'!I2:I{last},\"Medium\")", "Score 8 to 14; fortnightly review"),
        ("Low severity", f"=COUNTIF('Risk Register'!I2:I{last},\"Low\")", "Score 1 to 7; reviewed at phase gates"),
        ("Highest score on the register", f"=MAX('Risk Register'!H2:H{last})", "Drives the escalation cadence"),
        ("Mean score", f"=ROUND(AVERAGE('Risk Register'!H2:H{last}),1)", "Overall exposure indicator"),
        ("Open", f"=COUNTIF('Risk Register'!M2:M{last},\"Open\")", "Live and being managed"),
        ("Mitigated", f"=COUNTIF('Risk Register'!M2:M{last},\"Mitigated\")", "Control in place and evidenced"),
        ("Accepted", f"=COUNTIF('Risk Register'!J2:J{last},\"Accept\")", "Consciously carried, reasoning recorded"),
        ("Transferred", f"=COUNTIF('Risk Register'!J2:J{last},\"Transfer\")", "Carried by another party"),
        ("Compliance-category risks", f"=COUNTIF('Risk Register'!B2:B{last},\"Compliance\")", "Cross-check against AVSAR-DPD-009"),
        ("Risks with no target date", f"=COUNTIF('Risk Register'!N2:N{last},\"\")", "Must be zero"),
    ], note="Score and Severity on the register are formulas. Change a Likelihood or Impact cell and "
            "every figure here updates; nothing on this tab is typed.")

    wb.save(path)
    return path


def build_timeline(path):
    wb = new_workbook()

    cover_sheet(
        wb, "timeline",
        "Phase plan and week-level Gantt for the AVSAR programme, from policy mapping through the "
        "hackathon submission and on to deployment readiness. Phases marked Parallel overlap "
        "deliberately, which is what allows a five-stage mechanism to be built and tested inside a "
        "single quarter.",
        [("1.0", "03 Sep 2026", "TandSol", "Baseline. Sixteen phases; fourteen weeks to submission "
          "plus two quarters of deployment readiness.")],
        extra=[("Critical path", "P1 - P2 - P3 - P5 - P6 - P12 - P15. The design-system and "
                                 "client phases run alongside the API and are not on the critical path."),
               ("Milestones", "M1 RTM baselined (01 Aug) - M2 schema and workflow frozen (15 Aug) - "
                              "M3 API feature complete (08 Sep) - M4 client feature complete (12 Sep) - "
                              "M5 end-to-end suite green (17 Sep) - M6 submission (30 Sep)")],
    )

    ws = wb.create_sheet("Timeline")
    write_table(
        ws,
        ["Phase", "Phase name", "Activities", "Start", "End", "Duration (working days)", "Owner", "Mode"],
        [list(p) for p in PHASES],
        widths=[8, 34, 82, 13, 13, 13, 20, 12],
        wrap_cols=(3,),
    )
    add_status_validation(ws, "H", 2, len(PHASES) + 1, ["Sequential", "Parallel"])

    g = wb.create_sheet("Gantt")
    g.sheet_view.showGridLines = False
    g["A1"] = "AVSAR programme - week view"
    g["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)

    hdr = 3
    g.cell(row=hdr, column=1, value="Phase").font = body_font(bold=True, color="FFFFFF")
    g.cell(row=hdr, column=1).fill = fill(NAVY)
    g.cell(row=hdr, column=2, value="Phase name").font = body_font(bold=True, color="FFFFFF")
    g.cell(row=hdr, column=2).fill = fill(NAVY)
    for i, (wk, label) in enumerate(WEEKS, start=3):
        c = g.cell(row=hdr - 1, column=i, value=wk)
        c.font = body_font(bold=True, color=NAVY, size=9)
        c.alignment = Alignment(horizontal="center")
        c2 = g.cell(row=hdr, column=i, value=label)
        c2.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        c2.fill = fill(NAVY)
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
        g.column_dimensions[get_column_letter(i)].width = 11
    g.row_dimensions[hdr].height = 28

    for r, p in enumerate(PHASES, start=hdr + 1):
        pid, name, _, _, _, _, _, mode = p
        pc = g.cell(row=r, column=1, value=pid)
        pc.font = body_font(bold=True)
        pc.border = BOX
        nc = g.cell(row=r, column=2, value=name)
        nc.font = body_font()
        nc.border = BOX
        start, end = GANTT[pid]
        for wcol in range(1, len(WEEKS) + 1):
            cell = g.cell(row=r, column=wcol + 2)
            cell.border = BOX
            if start <= wcol <= end:
                cell.fill = fill(SAFFRON if mode == "Parallel" else NAVY)
    g.column_dimensions["A"].width = 8
    g.column_dimensions["B"].width = 38
    g.freeze_panes = "C4"

    legend_row = hdr + len(PHASES) + 2
    g.cell(row=legend_row, column=1, value="Legend").font = body_font(bold=True, color=NAVY)
    g.cell(row=legend_row + 1, column=1).fill = fill(NAVY)
    g.cell(row=legend_row + 1, column=2, value="Sequential - on the critical path").font = body_font()
    g.cell(row=legend_row + 2, column=1).fill = fill(SAFFRON)
    g.cell(row=legend_row + 2, column=2, value="Parallel - overlaps a preceding phase").font = body_font()

    n = len(PHASES)
    summary_sheet(wb, "Timeline - summary", [
        ("Phases planned", f"=COUNTA(Timeline!A2:A{n + 1})", "Total rows on the Timeline tab"),
        ("Sequential phases", f"=COUNTIF(Timeline!H2:H{n + 1},\"Sequential\")", "On the critical path"),
        ("Parallel phases", f"=COUNTIF(Timeline!H2:H{n + 1},\"Parallel\")", "Overlap a preceding phase"),
        ("Total effort (working days)", f"=SUM(Timeline!F2:F{n + 1})", "Sum of phase durations, before overlap"),
        ("Effort to hackathon submission", f"=SUM(Timeline!F2:F{n})", "Excludes the deployment-readiness phase"),
        ("Longest phase (days)", f"=MAX(Timeline!F2:F{n + 1})", "Deployment readiness, spanning two quarters"),
        ("Mean phase length (days)", f"=ROUND(AVERAGE(Timeline!F2:F{n + 1}),1)", "Indicator of phase granularity"),
        ("Parallelisation share", f"=ROUND(COUNTIF(Timeline!H2:H{n + 1},\"Parallel\")/COUNTA(Timeline!A2:A{n + 1}),3)", "Format as a percentage"),
        ("Distinct owners", "=SUMPRODUCT((Timeline!G2:G17<>\"\")/COUNTIF(Timeline!G2:G17,Timeline!G2:G17&\"\"))", "Distinct values in the Owner column"),
    ], note="Total effort is the sum of phase durations and exceeds elapsed calendar time because "
            "nine phases run in parallel. See the Gantt tab for the elapsed view.")

    wb.save(path)
    return path
